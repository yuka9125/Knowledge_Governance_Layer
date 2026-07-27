#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
レビュー済みExcelから承認済みKnowledgeを出力する。

出力はJSON（serving が参照する契約ファイル）を基本としつつ、同じ内容のCSVも
併せて出力する。upsertのベースはJSON/CSVのどちらでも読み込める。
"""

from __future__ import annotations

import argparse
import re
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import load_workbook

from knowledge_distillation.approved_knowledge_io import (
    CSV_FORMAT,
    JSON_FORMAT,
    companion_path,
    detect_format,
    read_approved_knowledge,
    resolve_existing_path,
    write_approved_knowledge_both,
)


SHEET_NAME = "最終ナレッジ候補一覧"
APPROVED_REVIEW_RESULT = "採用"
JST = timezone(timedelta(hours=9))
SIMPLE_KNOWLEDGE_ID_PATTERN = re.compile(r"^k-(\d+)$")
DATED_KNOWLEDGE_ID_PATTERN = re.compile(r"^k-(\d{8})-(\d+)$")


def _cell_text(value: Any) -> str:
    """Excelセル値をJSON向け文字列に整える。"""
    if value is None:
        return ""
    return str(value).strip()


def _build_header_index(ws) -> Dict[str, int]:
    """1行目のヘッダー名から列番号へのマップを作る。"""
    headers: Dict[str, int] = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        header = _cell_text(cell.value)
        if header:
            headers[header] = col_idx
    return headers


def _required_value(ws, row_idx: int, headers: Dict[str, int], header: str) -> str:
    """必須列の値を取得する。"""
    if header not in headers:
        raise ValueError(f"必須列が見つかりません: {header}")
    return _cell_text(ws.cell(row=row_idx, column=headers[header]).value)


def _optional_value(ws, row_idx: int, headers: Dict[str, int], header: str) -> str:
    """任意列の値を取得する（列が無ければ空文字）。"""
    if header not in headers:
        return ""
    return _cell_text(ws.cell(row=row_idx, column=headers[header]).value)


EXISTING_FAQ_ID_HEADER = "既存FAQ_ID"
LINK_NAME_HEADERS = ("参考リンク・資料名", "リンク名")


def _optional_value_any(
    ws,
    row_idx: int,
    headers: Dict[str, int],
    header_candidates: List[str] | tuple[str, ...],
) -> str:
    """任意列を候補ヘッダー名から取得する（見つからなければ空文字）。"""
    for header in header_candidates:
        if header in headers:
            return _optional_value(ws, row_idx, headers, header)
    return ""


def _approved_date_stamp(approved_at_value: str) -> str:
    """approved_at から YYYYMMDD を取り出す。"""
    normalized = approved_at_value.strip()
    if normalized.endswith("Z"):
        normalized = normalized[:-1] + "+00:00"
    try:
        dt = datetime.fromisoformat(normalized)
    except ValueError:
        return datetime.now(JST).strftime("%Y%m%d")

    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=JST)
    return dt.astimezone(JST).strftime("%Y%m%d")


def _dated_knowledge_id_for_new(knowledge_id: str, date_stamp: str) -> str:
    """新規行の k-連番IDだけを k-YYYYMMDD-連番 へ変換する。"""
    if DATED_KNOWLEDGE_ID_PATTERN.match(knowledge_id):
        return knowledge_id

    match = SIMPLE_KNOWLEDGE_ID_PATTERN.match(knowledge_id)
    if not match:
        return knowledge_id
    return f"k-{date_stamp}-{match.group(1)}"


def load_approved_knowledge_from_excel(
    excel_path_or_file: str | Path | Any,
    approved_at: str | None = None,
    use_existing_faq_id: bool = False,
) -> List[Dict[str, str]]:
    """レビュー結果が「採用」の行だけ approved_knowledge 形式で返す。

    use_existing_faq_id=True のとき、既存FAQ_ID が入っている行（＝既存FAQの更新）は
    knowledge_id を既存FAQ_ID に揃える。後段の upsert マージで既存エントリに上書きされる。
    """
    wb = load_workbook(excel_path_or_file, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"シートが見つかりません: {SHEET_NAME}")

    ws = wb[SHEET_NAME]
    headers = _build_header_index(ws)
    approved_at_value = approved_at or datetime.now(JST).isoformat(timespec="seconds")
    approved_date = _approved_date_stamp(approved_at_value)

    approved_items: List[Dict[str, str]] = []
    for row_idx in range(2, ws.max_row + 1):
        review_result = _required_value(ws, row_idx, headers, "レビュー結果")
        if review_result != APPROVED_REVIEW_RESULT:
            continue

        knowledge_id = _required_value(ws, row_idx, headers, "ナレッジID")
        question = _required_value(ws, row_idx, headers, "候補_質問")
        answer = _required_value(ws, row_idx, headers, "候補_回答")
        category = _required_value(ws, row_idx, headers, "カテゴリ")
        link_names = _optional_value_any(ws, row_idx, headers, LINK_NAME_HEADERS)
        if not knowledge_id or not question or not answer:
            continue

        if use_existing_faq_id:
            existing_faq_id = _optional_value(
                ws, row_idx, headers, EXISTING_FAQ_ID_HEADER
            )
            if existing_faq_id:
                # 既存FAQの更新 → 既存IDに揃えて上書き対象にする
                knowledge_id = existing_faq_id
            else:
                knowledge_id = _dated_knowledge_id_for_new(
                    knowledge_id=knowledge_id,
                    date_stamp=approved_date,
                )

        approved_items.append(
            {
                "knowledge_id": knowledge_id,
                "question": question,
                "answer": answer,
                "category": category,
                "link_names": link_names,
                "approved_status": "approved",
                "approved_at": approved_at_value,
            }
        )

    return approved_items


def merge_approved_knowledge(
    base_items: List[Dict[str, str]],
    incoming_items: List[Dict[str, str]],
) -> List[Dict[str, str]]:
    """knowledge_id をキーに upsert マージする。

    - incoming が base と同じ knowledge_id を持てば**上書き**（既存FAQ更新）
    - 新しい knowledge_id は**末尾に追加**（新規）
    - base の並び順を保ち、新規は出現順で後ろに足す
    """
    merged: Dict[str, Dict[str, str]] = {}
    order: List[str] = []
    for item in list(base_items) + list(incoming_items):
        kid = str(item.get("knowledge_id", "")).strip()
        if not kid:
            continue
        if kid not in merged:
            order.append(kid)
        merged[kid] = item  # 後勝ち（incoming が base を上書き）
    return [merged[kid] for kid in order]


def _read_base_items(path: str | Path) -> List[Dict[str, str]]:
    """upsertのベースを読む。指定が無ければ同名の別フォーマットも探す。"""
    resolved = resolve_existing_path(path)
    if resolved is None:
        return []
    return read_approved_knowledge(resolved)


def export_approved_knowledge_files(
    excel_path_or_file: str | Path | Any,
    output_path: str | Path,
    approved_at: str | None = None,
    base_path: str | Path | None = None,
    merge: bool = True,
    csv_output_path: str | Path | None = None,
) -> Dict[str, Path]:
    """レビュー済みExcelから approved_knowledge を JSON と CSV で出力する。

    戻り値は {"json": Path, "csv": Path}。
    output_path は .json / .csv のどちらでもよく、指定しなかった側は同名の
    別拡張子に出力する（内容は同じ）。

    merge=True（既定）：既存の approved_knowledge（base_path、未指定なら output_path。
    JSON/CSVのどちらでも読める）に対して knowledge_id で upsert する。
    merge=False：採用行のスナップショットで上書きする。
    """
    approved_items = load_approved_knowledge_from_excel(
        excel_path_or_file=excel_path_or_file,
        approved_at=approved_at,
        use_existing_faq_id=merge,
    )

    if merge:
        base_source = base_path if base_path is not None else output_path
        base_items = _read_base_items(base_source)
        result_items = merge_approved_knowledge(base_items, approved_items)
    else:
        result_items = approved_items

    return write_approved_knowledge_both(
        items=result_items,
        path=output_path,
        csv_path=csv_output_path,
    )


def export_approved_knowledge_from_excel(
    excel_path_or_file: str | Path | Any,
    output_path: str | Path,
    approved_at: str | None = None,
    base_path: str | Path | None = None,
    merge: bool = True,
    csv_output_path: str | Path | None = None,
) -> Path:
    """レビュー済みExcelから approved_knowledge を出力する。

    merge=True（既定）：既存の approved_knowledge（base_path、未指定なら output_path。
    JSON/CSVのどちらでも可）に対して knowledge_id で upsert する。既存FAQの更新は
    既存IDに上書き、新規は追加するため重複が出にくい。
    merge=False：従来どおり「採用行のスナップショット」を上書き出力する。

    JSONとCSVの両方を出力し、戻り値は output_path のフォーマット側のパス。
    もう一方のパスは `approved_knowledge_companion_path()` で取得できる。
    """
    outputs = export_approved_knowledge_files(
        excel_path_or_file=excel_path_or_file,
        output_path=output_path,
        approved_at=approved_at,
        base_path=base_path,
        merge=merge,
        csv_output_path=csv_output_path,
    )
    return outputs[detect_format(output_path)]


def approved_knowledge_companion_path(path: str | Path) -> Path:
    """JSON出力に対するCSV出力（またはその逆）のパスを返す。"""
    return companion_path(path)


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "レビュー済みExcelから approved_knowledge を出力します"
            "（JSONとCSVの両方）。"
        )
    )
    parser.add_argument("excel_path", help="FAQ_final_result.xlsx のパス")
    parser.add_argument(
        "-o",
        "--output",
        default="data/approved_knowledge.json",
        help="出力先パス（.json / .csv）",
    )
    parser.add_argument(
        "--csv-output",
        default=None,
        help="CSV出力先（未指定なら--outputと同名の.csv）",
    )
    parser.add_argument(
        "--base",
        default=None,
        help=(
            "upsertのベースとする既存ファイル（JSON/CSV可。"
            "未指定なら--outputを使う）"
        ),
    )
    parser.add_argument(
        "--no-merge",
        dest="merge",
        action="store_false",
        help="upsertせず採用行のスナップショットで上書きする",
    )
    args = parser.parse_args()

    outputs = export_approved_knowledge_files(
        excel_path_or_file=args.excel_path,
        output_path=args.output,
        base_path=args.base,
        merge=args.merge,
        csv_output_path=args.csv_output,
    )
    print(f"approved_knowledge出力(JSON): {outputs[JSON_FORMAT]}")
    print(f"approved_knowledge出力(CSV): {outputs[CSV_FORMAT]}")


if __name__ == "__main__":
    main()
